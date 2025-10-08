import re
import pandas as pd
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import load_workbook

data = []
testas = []

def olbg_get():
    global data, testas

    driver = webdriver.Chrome()
    driver.get("https://www.olbg.com/betting-tips/Football/1")
    sleep(5)  # wait a bit longer for content

    # Find matches (using XPath because CSS fails on min-h-[84px])
    matches = driver.find_elements(By.XPATH, "//li[contains(@class,'min-h-')]")
    print(f"Found {len(matches)} matches")

    for match in matches:
        try:
            fixture = match.find_element(By.CSS_SELECTOR, "h5[itemprop='name']").text.strip()
            pick = match.find_element(By.CSS_SELECTOR, "h4").text.strip()
            competition = match.find_element(By.CSS_SELECTOR, "p.text-sm.truncate").text.strip()
            match_time = match.find_element(By.TAG_NAME, "time").get_attribute("datetime")
            win_info = match.find_element(By.CSS_SELECTOR, "b.text-xs.truncate").text.strip()

            try:
                conf_style = match.find_element(By.CSS_SELECTOR, "div[style*='--confidence']").get_attribute("style")
                confidence = re.search(r"(\d+)%", conf_style).group(1) if conf_style else ""
            except:
                confidence = ""

            try:
                comments = match.find_element(By.CSS_SELECTOR, "span.text-xs.flex").text.strip()
            except:
                comments = "0"

            data.append([fixture, pick, competition, match_time, win_info, confidence, comments])

        except Exception as e:
            print("Skipping match:", e)

    # Only save if data is collected
    if data:
        df = pd.DataFrame(data, columns=[
            "Fixture", "Pick", "Competition", "Time", "Win Info", "Confidence %", "Comments"
        ])
        save_name = "olbg_fixtures.xlsx"
        df.to_excel(save_name, index=False)

        wb = load_workbook(save_name)
        ws = wb.active

        last_row = ws.max_row
        ws[f"H{last_row+1}"] = f"=AVERAGE(F2:F{last_row})"

        testas.append(ws[f"H{last_row+1}"].value)
        wb.save(save_name)
        print(f"✅ Saved {len(data)} matches to {save_name}")
    else:
        print("⚠️ No matches found!")

    driver.quit()
