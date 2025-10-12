from selenium import webdriver
import os
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from datetime import datetime
import pickle
import re
import math
from time import sleep
from pathlib import Path
import undetected_chromedriver as uc
from openpyxl import load_workbook
today = datetime.now().strftime("%d")
options = Options()
options.add_argument("--start-maximized")
driver = uc.Chrome(options=options)
counts = 0
data = []
testas = []
today_folder = datetime.now().strftime("%Y-%m-%d")

def get_save_path(source_name):
    os.makedirs(today_folder, exist_ok=True)
    return os.path.join(today_folder, f"{source_name}_fixtures.xlsx")


def ai_goalies_cookies():
    try:
            button = driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div/div[2]/div/button[2]")
            button.click()
            counts += 1
    except Exception:
            pass
def ai_goalie_get_past(i: int):
    global counts
    data = []
    testas = []

    driver.get(f"https://ai-goalie.com/{i}.09.2025.html")

    # Load cookies
    try:
        cookies = pickle.load(open("cookies.pkl", "rb"))
        for cookie in cookies:
            driver.add_cookie(cookie)
        driver.refresh()
        sleep(1)
    except FileNotFoundError:
        print("No cookies.pkl found")
        return

    # Accept cookies button (only once)
    if counts == 0:
        try:
            button = driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div/div[2]/div/button[2]")
            button.click()
            counts += 1
        except Exception:
            pass

    rows = driver.find_elements(By.CSS_SELECTOR, "tr.match")

    for row in rows:
        try:
            # Date
            match_date = row.find_element(By.CSS_SELECTOR, "td.date").text.strip()

            # Fixture
            home = row.find_element(By.CSS_SELECTOR, ".home-team").text.strip()
            score = row.find_element(By.CSS_SELECTOR, ".score").text.strip()
            away = row.find_element(By.CSS_SELECTOR, ".away-team").text.strip()
            fixture = f"{home} - {away}: {score}"

            # Pick
            try:
                pick = row.find_element(By.CSS_SELECTOR, ".home-team u, .away-team u").text.strip()
            except Exception:
                pick = ""

            # Win %
            win_percent = row.find_elements(By.TAG_NAME, "td")[4].text.strip()

            # Result & total goals
            result = row.find_element(By.CSS_SELECTOR, "td.correct").text.strip()
            if result != "":
                parts = score.split(":")
                total = int(parts[0]) + int(parts[1])
            else:
                total = None

            # Expected Goals
            expected_goals = row.find_element(By.CSS_SELECTOR, "td.gp").text.strip()
            expected_goals = re.sub(r"[^\d.]", "", expected_goals)

            if expected_goals:
                expected_goals = float(expected_goals)
                goals_pick = math.ceil(expected_goals) + 0.5
                under = None if total is None else goals_pick > total
            else:
                goals_pick = None
                under = None

            # Append data if correct day
            parts = match_date.split()
            if parts:
                try:
                    day = int(parts[-1])
                    if day == i:
                        data.append([
                            match_date, fixture, expected_goals, pick,
                            goals_pick, win_percent, result, total, under
                        ])
                except ValueError:
                    continue
        except Exception as e:
            print("Skipping row due to error:", e)

    # Filter by Win % >= 60
    data_clean = []
    for row in data:
        try:
            win_val = float(row[5].replace("%", "").strip())
            if win_val >= 60:
                data_clean.append(row)
        except Exception:
            continue

    # Save to Excel
    df = pd.DataFrame(data_clean,
                      columns=["Date", "Fixture", "XG", "Pick", "Goals_Pick",
                               "Win %", "Result", "Total", "Under"])
    save_name = get_save_path(f"{today}_ai")
    df.to_excel(save_name, index=False)

    # Open with openpyxl to add formulas
    wb = load_workbook(save_name)
    ws = wb.active

    # Find last non-empty row in column I
    last_row = 0
    for r in range(1, ws.max_row + 1):
        if ws[f"I{r}"].value is not None:
            last_row = r

    # Add formulas
    ws[f"K{last_row + 1}"] = "=COUNTIF(I:I,TRUE)"
    ws[f"K{last_row + 2}"] = "=COUNTIF(I:I,FALSE)"
    ws[f"K{last_row + 3}"] = f"=K{last_row + 1}+K{last_row + 2}"
    ws[f"L{last_row + 1}"] = f"=(K{last_row + 1}/K{last_row + 3})*100"

    # Append percentage formula
    testas.append(ws[f"L{last_row + 1}"].value)

    wb.save(save_name)
def ai_goalie_get(today):
    data = []
    testas = []

    driver.get(f"https://ai-goalie.com/{today}.10.2025.html")

    # Load cookies

    # Accept cookies button (only once)
    try:
            wait = WebDriverWait(driver, 5)
            button = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/div/div/div[2]/div/button[2]")))
            button.click()
            counts += 1
    except Exception:
            pass

    rows = driver.find_elements(By.CSS_SELECTOR, "tr.match")

    for row in rows:
        try:
            # Date
            match_date = row.find_element(By.CSS_SELECTOR, "td.date").text.strip()

            # Fixture
            home = row.find_element(By.CSS_SELECTOR, ".home-team").text.strip()
            score = row.find_element(By.CSS_SELECTOR, ".score").text.strip()
            away = row.find_element(By.CSS_SELECTOR, ".away-team").text.strip()
            fixture = f"{home} - {away}: {score}"

            # Pick
            try:
                pick = row.find_element(By.CSS_SELECTOR, ".home-team u, .away-team u").text.strip()
            except Exception:
                pick = ""

            # Win %
            win_percent = row.find_elements(By.TAG_NAME, "td")[4].text.strip()

            # Result & total goals
            result = ""
            # result_w = row.find_element(By.CSS_SELECTOR, "span.result-indicator.result-w")
            # if result_w:
            #     win = True
            # elif len(result_w)==0:
            #     result_l = row.find_element(By.CSS_SELECTOR, "span.result-indicator.result-l")
            #     win = False
            # if win:
            #     result = "Y"
            # else:
            #     result = "X"
            # if win == True or win == False:
            #     parts = score.split(":")
            #     total = int(parts[0]) + int(parts[1])
            # else:
            #     total = None
            parts = score.split(":")
            total = int(parts[0]) + int(parts[1])

            # Expected Goals
            expected_goals = row.find_element(By.CSS_SELECTOR, "td.gp").text.strip()
            expected_goals = re.sub(r"[^\d.]", "", expected_goals)

            if expected_goals:
                expected_goals = float(expected_goals)
                goals_pick = math.ceil(expected_goals) + 0.5
                under = None if total is None else goals_pick > total
            else:
                goals_pick = None
                under = None

            # Append data if correct day
            parts = match_date.split()
            if parts:
                try:
                    day = int(parts[-1])
                    data.append([
                            match_date, fixture, expected_goals, pick,
                            goals_pick, win_percent, result, total, under
                        ])
                except ValueError:
                    continue
        except Exception as e:
            continue
            # print("Skipping row due to error:", e)
    pf = pd.DataFrame(data,
                      columns=["Date", "Fixture", "XG", "Pick", "Goals_Pick",
                               "Win %", "Result", "Total", "Under"])
    save_name = get_save_path(f"{today}_ai.full")
    pf.to_excel(save_name, index=False)
    # Filter by Win % >= 60
    data_clean = []
    for row in data:
        try:
            win_val = float(row[5].replace("%", "").strip())
            if win_val >= 55:
                data_clean.append(row)
        except Exception:
            continue

    # Save to Excel
    df = pd.DataFrame(data_clean,
                      columns=["Date", "Fixture", "XG", "Pick", "Goals_Pick",
                               "Win %", "Result", "Total", "Under"])
    save_name = get_save_path(f"{today}")
    df.to_excel(save_name, index=False)

    # Open with openpyxl to add formulas
    wb = load_workbook(save_name)
    ws = wb.active

    # Find last non-empty row in column I
    last_row = 0
    for r in range(1, ws.max_row + 1):
        if ws[f"I{r}"].value is not None:
            last_row = r

    # Add formulas
    ws[f"K{last_row + 1}"] = "=COUNTIF(I:I,TRUE)"
    ws[f"K{last_row + 2}"] = "=COUNTIF(I:I,FALSE)"
    ws[f"K{last_row + 3}"] = f"=K{last_row + 1}+K{last_row + 2}"
    ws[f"L{last_row + 1}"] = f"=(K{last_row + 1}/K{last_row + 3})*100"

    # Append percentage formula
    testas.append(ws[f"L{last_row + 1}"].value)

    wb.save(save_name)
def oddspedia_get(today):
    """
    Scrapes football betting tips from Oddspedia using the globally defined driver.
    Filters for consensus tips with 60% confidence or higher.
    """
    data = []
    testas = []
    wait = WebDriverWait(driver, 10)

    try:
        driver.get("https://oddspedia.com/football/tips")
        sleep(5) # Initial wait for page load

        # Step 1: Click "By Consensus" tab using the provided XPath
        try:
            # Absolute XPath provided by user
            CONSENSUS_BUTTON_XPATH = "/html/body/div[1]/div[2]/div/div[1]/div[2]/div[2]/div[2]/div/main/div[3]/div[1]/div[2]/ul/li[2]/button"
            wait = WebDriverWait(driver, 5)
            consensus_tab = wait.until(
                EC.element_to_be_clickable((By.XPATH, CONSENSUS_BUTTON_XPATH))
            )
            consensus_tab.click()
            sleep(5)  # Wait for content to reload
        except Exception as e:
            print("Failed to click 'By Consensus' button with provided XPath. Skipping sorting by consensus:", e)

        # Step 2: Open sort dropdown and select "Tips Amount"
        try:
            # Click the sort dropdown toggle
            sort_toggle = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".tips-sort-options__dropdown .old-dropdown__toggle"))
            )
            sort_toggle.click()
            sleep(5)

            # Click "Tips Amount" option
            tips_amount_option = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'old-dropdown__list-item') and contains(., 'Tips Amount')]"))
            )
            tips_amount_option.click()
            sleep(5)  # Allow UI to update
            odds_option = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='breadcrumb-bar']/div/div[2]/ul/li[2]/div/button/span[2]"))
            )
            odds_option.click()
            sleep(1)
            eu_odds = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='breadcrumb-bar']/div/div[2]/ul/li[2]/div/div/div[1]"))
            )
            eu_odds.click()
            
        except Exception as e:
            print("Failed to sort by 'Tips Amount'. Scraping current list order:", e)

       # Step 3: Scrape the matches
        matches = driver.find_elements(By.CSS_SELECTOR, "div.tip-by-consensus")

        # Define the exclusion keywords
        EXCLUSION_KEYWORDS = ["Yes", "Over", "Under", "-", "+","Draw"]

        for match in matches:
            try:
                # 1. Competition
                competition = match.find_element(
                    By.CSS_SELECTOR, "li.old-match-breadcrumbs__item div.masked-url"
                ).get_attribute("title").strip()

                # 2. Fixture
                home = match.find_element(By.CSS_SELECTOR, ".match-teams .match-team:nth-child(1) .match-team__name").text.strip()
                away = match.find_element(By.CSS_SELECTOR, ".match-teams .match-team:nth-child(2) .match-team__name").text.strip()
                fixture = f"{home} vs {away}"

                # 3. Pick
                pick = match.find_element(By.CSS_SELECTOR, ".tip-by-consensus__meta").text.strip().replace("Full Time Result:", "").strip()
                
                # 3. Odds
                odds = match.find_element(By.CSS_SELECTOR, "span.odd__value").text.strip()
                # 🛑 NEW FILTER: Check if pick contains any exclusion keywords
                skip_match = False
                for keyword in EXCLUSION_KEYWORDS:
                    if keyword in pick:
                        skip_match = True
                        break
                
                if skip_match:
                    continue  # Skip this match and move to the next one

                # 4. Match Time
                match_time = ""
                try:
                    time_elem = match.find_element(By.CSS_SELECTOR, ".match-date__time")
                    match_time = " ".join(time_elem.text.split()).replace(' ', ' ')
                except:
                    pass

                # 5. Win Info (Tip Amounts, e.g., "52 of 89 Tips")
                tip_amount_info = match.find_element(By.CSS_SELECTOR, ".tip-by-consensus__bar__meta").text.strip()
                win_info = tip_amount_info 
                
                # 6. Confidence
                confidence_text = match.find_element(By.CSS_SELECTOR, ".old-progress-bar__value").text.strip()
                confidence = confidence_text.replace("%", "")


                # Filter 2: Only include if confidence >= 60%
                if int(confidence) >= 60:
                    data.append([fixture, pick, competition, match_time, win_info, confidence,odds])

            except Exception as e:
                # print(f"Skipping match due to error: {e}")
                continue

        # Step 4: Save to Excel
        if data:
            df = pd.DataFrame(data, columns=["Fixture", "Pick", "Competition", "Time", "Win Info", "Confidence %","Odds"])
            save_name = get_save_path(f"{today}_oddspedia")
            df.to_excel(save_name, index=False)

            wb = load_workbook(save_name)
            ws = wb.active
            last_row = ws.max_row
            
            # Add Average Confidence formula
            if last_row >= 2:
                ws[f"H{last_row + 1}"] = f"=AVERAGE(F2:F{last_row})"
                testas.append(ws[f"H{last_row + 1}"].value) 
            
            wb.save(save_name)

    except Exception as e:
        print(f"A major error occurred during Oddspedia scraping: {e}")
    
    # The global driver is NOT quit here. It must be quit at the very end of the script.
def olbg_get(today):
    data = []
    testas = []
    driver.get("https://www.olbg.com/betting-tips/Football/1")
    sleep(5)  # wait a bit longer for content

    # Find matches (using XPath because CSS fails on min-h-[84px])
    matches = driver.find_elements(By.XPATH, "//li[contains(@class,'min-h-')]")
    print(f"Found {len(matches)} matches")

    for match in matches:
        try:
            fixture = match.find_element(By.CSS_SELECTOR, "h5[itemprop='name']").text.strip()
            pick = match.find_element(By.CSS_SELECTOR, "h4").text.strip()
            competition = match.find_element(By.CSS_SELECTOR, "p.text-sm.truncate").text.strip()
            match_time = match.find_element(By.TAG_NAME, "time").get_attribute("datetime")
            win_info = match.find_element(By.CSS_SELECTOR, "b.text-xs.truncate").text.strip()

            try:
                conf_style = match.find_element(By.CSS_SELECTOR, "div[style*='--confidence']").get_attribute("style")
                confidence = re.search(r"(\d+)%", conf_style).group(1) if conf_style else ""
            except:
                confidence = ""

            try:
                comments = match.find_element(By.CSS_SELECTOR, "span.text-xs.flex").text.strip()
            except:
                comments = "0"

            data.append([fixture, pick, competition, match_time, win_info, confidence, comments])

        except Exception as e:
            print("Skipping match:", e)

    # Only save if data is collected
    if data:
        df = pd.DataFrame(data, columns=[
            "Fixture", "Pick", "Competition", "Time", "Win Info", "Confidence %", "Comments"
        ])
        save_name = get_save_path(f"{today}_olbg")
        df.to_excel(save_name, index=False)

        wb = load_workbook(save_name)
        ws = wb.active

        last_row = ws.max_row
        ws[f"H{last_row+1}"] = f"=AVERAGE(F2:F{last_row})"

        testas.append(ws[f"H{last_row+1}"].value)
        wb.save(save_name)
        print(f"✅ Saved {len(data)} matches to {save_name}")
    else:
        print("⚠️ No matches found!")
SCRIPT_DIR = Path(__file__).parent.resolve()

def compare_confidence_sources(ai_goalie_file, olbg_file, oddspedia_file):
    
    # --- Helper function for word-based name cleaning and tokenization ---
    def get_match_tokens(name):
        """Extracts significant, clean, lowercase words for comparison."""
        if pd.isna(name):
            return set()
        name = str(name).lower()
        # Remove common, noisy words that often appear in picks/fixtures
        name = re.sub(r'\b(fc|utd|united|city|ac|cf|sc|tsv|sv|fk|sk|draw|the|and|or|of|a|an)\b', ' ', name)
        # Remove non-alphanumeric characters, then split into words
        words = re.sub(r'[^a-z0-9\s]', ' ', name).split()
        # Filter out very short words that are likely generic
        return {word for word in words if len(word) > 2}
    # -----------------------------------------------------------------
    today_path = SCRIPT_DIR / today_folder
    # 1. Load DataFrames
    try:
        df_ai = pd.read_excel(today_path / ai_goalie_file)
        df_olbg = pd.read_excel(today_path / olbg_file)
        df_oddspedia = pd.read_excel(today_path / oddspedia_file)
    except FileNotFoundError as e:
        print(f"Error: One or more required Excel files not found: {e}")
        return pd.DataFrame()
    except Exception as e:
        print(f"Error loading Excel files: {e}")
        return pd.DataFrame()

    # 2. Prepare Match Tokens for all 'Pick' columns
    df_ai['Match_Tokens'] = df_ai['Pick'].apply(get_match_tokens)
    df_olbg['Pick_Text'] = df_olbg['Pick'].astype(str).str.lower()
    df_oddspedia['Pick_Text'] = df_oddspedia['Pick'].astype(str).str.lower()
    
    df_ai.dropna(subset=['Match_Tokens'], inplace=True)
    comparison_data = []

    # 3. Iterate and Compare using Word Match
    for index, row in df_ai.iterrows():
        ai_tokens = row['Match_Tokens']
        ai_original_pick = row['Pick']
        ai_fixture = row['Fixture']
        ai_result = row['Result']
        
        
        # Clean AI Goalie confidence
        ai_confidence_str = str(row['Win %']).replace('%', '').strip()
        olbg_confidence = None
        oddspedia_confidence = None
        result = None
        odds = None
        # if df_ai[]
        # --- Check OLBG ---
        # Find OLBG rows where the OLBG Pick Text contains any word from the AI Goalie tokens
        olbg_match = df_olbg[
            df_olbg['Pick_Text'].apply(lambda x: any(token in x for token in ai_tokens))
        ]
        if ai_result:
            result = ai_result
        if not olbg_match.empty:
            # Found a match: take the confidence (first one if multiple exist)
            olbg_confidence = olbg_match.iloc[0]['Confidence %']

        # --- Check Oddspedia ---
        # Find Oddspedia rows where the Oddspedia Pick Text contains any word from the AI Goalie tokens
        oddspedia_match = df_oddspedia[
            df_oddspedia['Pick_Text'].apply(lambda x: any(token in x for token in ai_tokens))
        ]
        if not oddspedia_match.empty:
            oddspedia_confidence = oddspedia_match.iloc[0]['Confidence %']
            odds = oddspedia_match.iloc[0]['Odds']
        
        # 4. Append Result (Only if at least one other source has a matching pick)
        if olbg_confidence is not None or oddspedia_confidence is not None:
            comparison_data.append({
                "Fixture": ai_fixture,
                "Pick": ai_original_pick,
                "AI_Confidence": ai_confidence_str,
                "OLBG_Confidence": olbg_confidence,
                "Oddspedia_Confidence": oddspedia_confidence,
                "Odds": odds,
                "Result": result
            })

    # 5. Create Final DataFrame and Save
    df_comparison = pd.DataFrame(comparison_data)
    
    if df_comparison.empty:
        save_name = f"{today}_combined_confidence"
        save_name = os.path.join(today_folder, f"{save_name}.xlsx")
        df_comparison.to_excel(save_name, index=False)
        print("Found 0 common picks. No output file created.")
        return df_comparison

    # Clean up and convert confidence columns to numeric
    for col in ['AI_Confidence', 'OLBG_Confidence', 'Oddspedia_Confidence']:
        df_comparison[col] = df_comparison[col].astype(str).str.replace('%', '', regex=False).str.strip()
        df_comparison[col] = pd.to_numeric(df_comparison[col], errors='coerce')

    print(f"Found {len(df_comparison)} common picks.")
    # save_folder = Path("X:/Colab Notebooks")
    # file_name = f"{today}_combined_confidence_xlsx"
    # full_path = save_folder / file_name
   
    save_name = f"{today}_combined_confidence"
    save_name = os.path.join(today_folder, f"{save_name}.xlsx")
    df_comparison.to_excel(save_name, index=False)
    print(f"Results saved to {today}_combined_confidence.xlsx")
    
    return df_comparison

yesterday = int(today) -1
yesterday = f"0{str(yesterday)}"
# oddspedia_get(today)
compare_confidence_sources(f"{today}_fixtures.xlsx",f"{today}_olbg_fixtures.xlsx",f"{today}_oddspedia_fixtures.xlsx")

# compare_confidence_sources(f"{yesterday}_fixtures.xlsx","{today}_olbg_fixtures.xlsx","{today}_oddspedia_fixtures.xlsx")